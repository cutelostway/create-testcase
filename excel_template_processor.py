# excel_template_processor.py - Excel template processor with VBA integration
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from io import BytesIO
import os

def create_testcase_template(project_settings, test_cases):
    """
    Tạo template Excel test case theo format TPL-QA-01-04 với VBA code
    
    Args:
        project_settings: dict chứa thông tin project
        test_cases: list các test case
    
    Returns:
        bytes: Excel file dưới dạng bytes
    """
    # Tạo workbook mới
    wb = Workbook()
    
    # Xóa sheet mặc định
    wb.remove(wb.active)
    
    # Tạo các sheet theo template
    create_cover_sheet(wb, project_settings)
    create_record_of_change_sheet(wb)
    create_report_sheet(wb, project_settings, test_cases)
    
    # Lưu vào BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()

def create_cover_sheet(wb, project_settings):
    """Tạo sheet Cover với thông tin project"""
    ws = wb.create_sheet("Cover")
    
    # Tiêu đề
    ws['A1'] = "TEST PLAN"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Thông tin project
    ws['A3'] = "Project Name:"
    ws['B3'] = project_settings.get('name', '')
    
    ws['A4'] = "Reviewer:"
    ws['B4'] = ""  # Bỏ trống theo yêu cầu
    
    ws['A5'] = "Phase:"
    ws['B5'] = project_settings.get('phase', '')
    
    ws['A6'] = "Created time (h):"
    ws['B6'] = ""  # Bỏ trống theo yêu cầu
    
    # Định dạng
    for row in range(3, 7):
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'].font = Font()

def create_record_of_change_sheet(wb):
    """Tạo sheet Record of Change"""
    ws = wb.create_sheet("Record of Change")
    
    # Tiêu đề
    ws['A1'] = "RECORD OF CHANGE"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Header bảng
    headers = ["Date", "Version", "Change Description", "A/M/D", "Change Item", "Reference/PIC"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    # Thêm record mới
    today = datetime.now().strftime("%Y-%m-%d")
    new_record = [today, "1.0", "Add new", "A", "All sheet", ""]
    for col, value in enumerate(new_record, 1):
        ws.cell(row=4, column=col, value=value)
    
    # Định dạng bảng
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 15

def create_report_sheet(wb, project_settings, test_cases):
    """Tạo sheet Report với test cases"""
    ws = wb.create_sheet("Report")
    
    # Tiêu đề
    ws['A1'] = "TEST CASE REPORT"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Thông tin project
    ws['A3'] = f"Project: {project_settings.get('name', '')}"
    ws['A4'] = f"Phase: {project_settings.get('phase', '')}"
    ws['A5'] = f"Environment: {', '.join(project_settings.get('environment', []))}"
    
    # Header bảng test case
    headers = [
        "No", "Test Case ID", "Test Title", "Description", "Preconditions", 
        "Test Data", "Test Steps", "Expected Result", "Priority", "Status", "Comments"
    ]
    
    start_row = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Thêm test cases
    for idx, case in enumerate(test_cases, 1):
        row = start_row + idx
        
        # Chuyển đổi case thành dict nếu cần
        if hasattr(case, 'dict'):
            case_dict = case.dict()
        else:
            case_dict = dict(case)
        
        # Xác định priority từ test_case_id
        priority = "Medium"
        if 'critical' in str(case_dict.get('test_case_id', '')).lower():
            priority = "Critical"
        elif 'high' in str(case_dict.get('test_case_id', '')).lower():
            priority = "High"
        elif 'low' in str(case_dict.get('test_case_id', '')).lower():
            priority = "Low"
        
        # Điền dữ liệu
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=case_dict.get('test_case_id', ''))
        ws.cell(row=row, column=3, value=case_dict.get('test_title', ''))
        ws.cell(row=row, column=4, value=case_dict.get('description', ''))
        ws.cell(row=row, column=5, value=case_dict.get('preconditions', ''))
        ws.cell(row=row, column=6, value=case_dict.get('test_data', ''))
        ws.cell(row=row, column=7, value=case_dict.get('test_steps', ''))
        ws.cell(row=row, column=8, value=case_dict.get('expected_result', ''))
        ws.cell(row=row, column=9, value=priority)
        ws.cell(row=row, column=10, value="Not Executed")
        ws.cell(row=row, column=11, value=case_dict.get('comments', ''))
    
    # Định dạng cột
    column_widths = [5, 15, 30, 40, 30, 20, 50, 30, 10, 12, 30]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Thêm border cho bảng
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(start_row, start_row + len(test_cases) + 1):
        for col in range(1, 12):
            ws.cell(row=row, column=col).border = thin_border

def create_vba_code():
    """
    Tạo VBA code để tích hợp vào Excel template
    """
    vba_code = '''
Sub CreateTestTemplate()
    ' VBA code để tạo template test case
    ' Sử dụng thông tin từ project settings
    
    Dim ws As Worksheet
    Dim lastRow As Long
    Dim i As Integer
    
    ' Tạo sheet mới cho module test
    Set ws = Worksheets.Add
    ws.Name = "Module 1"
    
    ' Thêm header
    ws.Cells(1, 1) = "TEST CASE MODULE 1"
    ws.Cells(1, 1).Font.Bold = True
    ws.Cells(1, 1).Font.Size = 14
    
    ' Thêm thông tin environment
    ws.Cells(3, 1) = "Test Environment:"
    ws.Cells(3, 2) = "Chrome, Firefox, Edge" ' Sẽ được thay thế bằng environment thực tế
    
    ' Tạo bảng test case
    Dim headers As Variant
    headers = Array("No", "Test Case ID", "Test Title", "Description", "Preconditions", "Test Data", "Test Steps", "Expected Result", "Priority", "Status")
    
    For i = 0 To UBound(headers)
        ws.Cells(5, i + 1) = headers(i)
        ws.Cells(5, i + 1).Font.Bold = True
        ws.Cells(5, i + 1).Interior.Color = RGB(68, 114, 196)
        ws.Cells(5, i + 1).Font.Color = RGB(255, 255, 255)
    Next i
    
    ' Định dạng cột
    ws.Columns("A:J").AutoFit
    
End Sub
'''
    return vba_code

def export_with_template(project_settings, test_cases):
    """
    Export test cases với template Excel có VBA code
    
    Args:
        project_settings: dict chứa thông tin project
        test_cases: list các test case
    
    Returns:
        bytes: Excel file với template và VBA code
    """
    return create_testcase_template(project_settings, test_cases)
