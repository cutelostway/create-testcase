# template_updater.py - Cập nhật template Excel gốc với giá trị mới
import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from datetime import datetime
from io import BytesIO
import os

def update_template_with_values(project_settings, test_cases):
    """
    Chọn đúng template theo số lượng Environment (1→7) và cập nhật nội dung.

    - Template: TPL_TestResult_v1.0_{N}En.xlsm với N = số môi trường đã chọn (giới hạn 1..7)
    - Cover sheet: giữ nguyên logic hiện tại
    - Report sheet: chỉ điền tên Environment vào cột C, bắt đầu từ C4; các ô khác giữ nguyên
    - Các sheet khác: giữ nguyên (không chỉnh sửa)
    """
    environments = project_settings.get('environment', ['Chrome'])
    num_environments = max(1, min(7, len(environments)))

    template_path = f"TPL_TestResult_v1.0_{num_environments}En.xlsm"

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file không tồn tại: {template_path}")

    # Tạo bản copy trong memory
    buffer = BytesIO()

    # Copy file gốc
    with open(template_path, 'rb') as source_file:
        buffer.write(source_file.read())

    buffer.seek(0)

    # Load workbook từ buffer
    wb = load_workbook(buffer, keep_vba=True)

    # Cập nhật các sheet theo yêu cầu
    update_cover_sheet(wb, project_settings)
    update_report_sheet(wb, project_settings, test_cases)
    update_module_sheets(wb, project_settings, test_cases)

    # Lưu lại vào buffer mới
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    return output_buffer.read()

def update_cover_sheet(wb, project_settings):
    """Cập nhật sheet Cover với thông tin project - giữ nguyên logic hiện tại"""
    try:
        ws = wb["Cover"]
    except KeyError:
        # Nếu không có sheet Cover, tạo mới
        ws = wb.create_sheet("Cover")
    
    try:
        # Cập nhật các cell theo yêu cầu mới
        # Cell D3 = Project Name
        ws['D3'] = project_settings.get('name', '')
        
        # Cell D4 = Phase
        ws['D4'] = project_settings.get('phase', '')
        
        # Cell G3 = blank
        ws['G3'] = ""
        
        # Dòng 8 - Record of Change
        today = datetime.now().strftime("%Y-%m-%d")
        ws['B8'] = today  # Date
        ws['C8'] = "1.0"  # Version
        ws['D8'] = "Add new"  # Change Description
        ws['E8'] = "A"  # A/M/D
        ws['F8'] = "All sheet"  # Change Item
            
    except Exception as e:
        print(f"Lỗi khi cập nhật Cover sheet: {e}")

def update_report_sheet(wb, project_settings, test_cases):
    """
    Cập nhật sheet Report theo yêu cầu đơn giản:
    - Ghi tên môi trường (chọn ở create-project) vào cột C, bắt đầu C4
    - Mọi giá trị khác giữ nguyên (không xóa/merge/thêm dòng)
    """
    try:
        ws = wb["Report"]
    except KeyError:
        print("Không tìm thấy sheet Report")
        return

    try:
        environments = project_settings.get('environment', [])
        for idx, env in enumerate(environments):
            ws[f"C{4 + idx}"] = env
        # Không thay đổi gì thêm
        print(f"Report sheet: ghi {len(environments)} environment(s) vào C4..C{3+len(environments)}")
    except Exception as e:
        print(f"Lỗi khi cập nhật Report sheet: {e}")

def update_module_sheets(wb, project_settings, test_cases):
    """
    Cập nhật các sheet Module với dữ liệu test case:
    - x = số environment đã chọn
    - y = số test case sau khi generate
    - Điền dữ liệu từ dòng (6+x) đến dòng (6+x+y)
    """
    try:
        environments = project_settings.get('environment', ['Chrome'])
        x = len(environments)  # số environment
        y = len(test_cases)    # số test case
        
        # Lấy danh sách các sheet Module (bỏ qua Cover và Report)
        module_sheets = [name for name in wb.sheetnames if name not in ["Cover", "Report"]]
        
        for sheet_name in module_sheets:
            ws = wb[sheet_name]
            print(f"Cập nhật Module sheet: {sheet_name}")
            
            # Điền dữ liệu từ dòng (6+x) đến dòng (6+x+y)
            start_row = 6 + x
            end_row = 6 + x + y
            
            for i, test_case in enumerate(test_cases):
                current_row = start_row + i
                
                # Chuyển đổi test_case thành dict nếu cần
                if hasattr(test_case, 'dict'):
                    tc_dict = test_case.dict()
                elif hasattr(test_case, '__dict__'):
                    tc_dict = test_case.__dict__
                else:
                    tc_dict = test_case
                
                # Điền dữ liệu vào các cột
                ws[f"B{current_row}"] = tc_dict.get('description', '')
                ws[f"C{current_row}"] = tc_dict.get('preconditions', '')
                ws[f"D{current_row}"] = tc_dict.get('test_steps', '')
                ws[f"E{current_row}"] = tc_dict.get('test_data', '')
                ws[f"F{current_row}"] = tc_dict.get('expected_result', '')
                
                # Cột G: Priority dựa trên test_case_id
                priority = get_priority_from_test_case(tc_dict)
                ws[f"G{current_row}"] = priority
                
            print(f"  Đã điền {y} test cases vào dòng {start_row}-{end_row-1}")
            
    except Exception as e:
        print(f"Lỗi khi cập nhật Module sheets: {e}")

def get_priority_from_test_case(tc_dict):
    """
    Xác định priority dựa trên test_case_id và mapping theo yêu cầu:
    - Critical priority và High Priority → High
    - Medium Priority → Normal  
    - Low Priority → Low
    """
    test_case_id = str(tc_dict.get('test_case_id', '')).lower()
    
    if 'critical' in test_case_id:
        return 'High'  # Critical → High
    elif 'high' in test_case_id:
        return 'High'  # High → High
    elif 'medium' in test_case_id:
        return 'Normal'  # Medium → Normal
    elif 'low' in test_case_id:
        return 'Low'  # Low → Low
    else:
        return 'Normal'  # Default priority → Normal



# ===== HELPER FUNCTIONS FOR VBA CONVERSION =====

def find_cell_by_text(ws, text):
    """Tìm cell chứa text cụ thể và trả về row number"""
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and text.lower() in str(cell_value).lower():
                return row
    return None

def column_number_to_letter(column_number):
    """Chuyển đổi số cột thành chữ cái (A, B, C, ...)"""
    result = ""
    while column_number > 0:
        column_number -= 1
        result = chr(65 + column_number % 26) + result
        column_number //= 26
    return result

def check_merge_conflict(ws, start_row, end_row, start_col, end_col):
    """
    Kiểm tra xem có xung đột với merged cells hiện có không
    """
    for merged_range in ws.merged_cells.ranges:
        # Kiểm tra xem có overlap không
        if (merged_range.min_row <= end_row and merged_range.max_row >= start_row and
            merged_range.min_col <= end_col and merged_range.max_col >= start_col):
            return True
    return False

def safe_merge_cells(ws, merge_ranges):
    """
    Merge nhiều cells một cách an toàn, bỏ qua nếu có xung đột
    """
    for merge_range in merge_ranges:
        try:
            # Parse range để lấy thông tin
            start_cell, end_cell = merge_range.split(':')
            start_col = ord(start_cell[0]) - ord('A') + 1
            start_row = int(start_cell[1:])
            end_col = ord(end_cell[0]) - ord('A') + 1
            end_row = int(end_cell[1:])
            
            # Kiểm tra xung đột
            if not check_merge_conflict(ws, start_row, end_row, start_col, end_col):
                ws.merge_cells(merge_range)
                print(f"    ✅ Merged {merge_range}")
            else:
                print(f"    ⚠️ Bỏ qua {merge_range} do xung đột")
        except Exception as e:
            print(f"    ❌ Lỗi merge {merge_range}: {e}")


